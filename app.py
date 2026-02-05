import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from logic import process_uploaded_file, suggest_similar_columns
import re

# Cache the main processing function for faster repeated runs
@st.cache_data(show_spinner=False)
def cached_process_uploaded_file(input_df, selected_rules, thresholds, weights):
    return process_uploaded_file(input_df, selected_rules, thresholds, weights)

st.set_page_config(
    page_title="TQA",
    layout="wide",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': """
# 🎯 Ticket Quality Audit Tool V.1

## 📋 Overview
This advanced audit dashboard helps analyze and improve ticket quality across different towers, applications, and associates. Built for comprehensive quality assessment and performance tracking.

## ⚡ Key Features
• **Smart Template Generation** - Pre-configured Excel templates with sample data
• **Automated Quality Scoring** - 21+ validation criteria with customizable weights
• **Interactive Dashboards** - Real-time charts and performance metrics
• **Multi-dimensional Analysis** - Tower-wise, Application-wise, and Associate-wise insights
• **Color-coded Performance** - Visual indicators for quick quality assessment
• **Export Capabilities** - Download results in professional Excel format

## 📊 Validation Criteria
✅ Short & Long Description Quality  
✅ SLA Compliance (Response & Resolution)  
✅ Knowledge Base Utilization  
✅ Work Notes & Documentation  
✅ Priority & Assignment Validation  
✅ Security & PA Violation Checks  
✅ User Confirmation & Communication  

## 🎨 Chart Features
• **Associate Performance**: Dual metrics showing ticket count + average score
• **Tower Analysis**: Average audit scores by business unit
• **Application Insights**: Quality metrics by application/CI
• **Category Performance**: Pass rates across all validation criteria

## 🛠️ Built With
• **Streamlit** - Interactive web framework
• **Plotly** - Advanced data visualizations
• **Pandas** - Data processing and analysis
• **OpenPyXL** - Excel file generation

## 📞 Support
For technical support or feature requests, contact Rohit.AvinashPagar@cognizant.com
---
*Version 1.0 | Built for Quality Excellence*
        """
    }
)

# st.title("🎯 TQA Tool by ISmO V.2")
# st.markdown("##### 🚀 Built for Quality Excellence")
# st.markdown("")
st.markdown("""
    <h1 style='text-align: center; color:#0d6efd; font-size: 5em; font-weight: bold;'>
        🎯 TQA Tool by ISmO V.1
    </h1>
    <p style='text-align: center; font-size: 1.3em; color:#444;'>
        🚀 Empowering <b>Quality Excellence</b> | <i>Ticket Quality Audit</i>
    </p>
""", unsafe_allow_html=True)

# --- Step 1: Enhanced Template Download  (27 columns headers)
TEMPLATE_COLUMNS = [
    "Number", "Priority", "State", "Pending reason",
    "Application Name / CI", "Assignment Group", "Assigned to", "Knowledge Article Used",
    "Short description", "Comments and Work notes", "Additional comments",
    "Description", "Closed", "Opened", "Closed by", "Subcategory",
    "Category", "Response Time", "Response SLA", "Resolution SLA", "Reopened",
    "Related Record", "Reassignment count", "Impact", "Urgency", "Age",
    "Has Attachments"
]

@st.cache_data
def generate_enhanced_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Input Template"
 
    # Simple, robust header styling
    for col_num, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Set appropriate column width - not too wide to avoid scrolling
        ws.column_dimensions[cell.column_letter].width = 14
    
    # Add sample data row
    sample_data = [
        "INC647839/RITM3551234", "3-Medium", "Closed Complete", "",
        "SAP_Production_Environment", "AMS.DIT.Enterprise.Apps.Cognizant.L2", "John.Doe", "KB0048434",
        "User unable to access SAP application after password reset", 
        "15/01/2025 09:30:00 - John Doe (Work notes)\nHello User,\n\nWe acknowledge your request. We have started looking into the ticket and we will get back to you if we need any additional information.\n\nRegards,\nSAP Support Team\n\n16/01/2025 14:20:00 - John Doe (Work notes)\nHello User,\n\nIssue has been resolved. Password reset completed and user access verified.\n\nRegards,\nSAP Support Team", 
        "16/01/2025 15:45:00 - John Doe (Additional comments)\nKnowledge article KB0048434:\nSAP User Access Management\n\nPassword reset performed successfully and user access restored.",
        "User reported login failure after recent password policy update. Investigation revealed account lock due to multiple failed attempts. Performed password reset and verified successful login.", 
        "16/01/2025 15:45:00", "15/01/2025 09:15:00", "Jane.Smith",
        "User Access Management", "Software", "2.5", "Met", "Met", "0/1/2 or True,False/Yes,No,y,n",
        "CHG0123456", "0", "3-Medium", "3-Medium", "1.7", "TRUE/FALSE"
    ]

    
    for col_num, value in enumerate(sample_data, 1):
        if col_num <= len(TEMPLATE_COLUMNS):
            cell = ws.cell(row=2, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='top')
 
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions_ws.cell(row=1, column=1, value="🎯 Ticket Audit Tool - Instructions").font = Font(bold=True, size=16)
    
    instructions = [
    "🎯 Ticket Audit Tool – Validation Rules & Instructions",
    "",
    "──────────────────────────────────────────────────────────────",
    "📌 General Notes:",
    "• If a required column is missing, the corresponding validation will fail for all tickets.",
    "• Use the provided mapping files (Tower_Maping.xlsx, Category_Subcategory_Mapping.xlsx) in the same folder.",
    "• Do NOT rename mapping files or their headers.",
    "• Do NOT rename the column headers in the input template",
    "",
    "──────────────────────────────────────────────────────────────",
    "✅ Validation Rules: and required columns in input file",
    "Ticket Number",
    "   • Required column: Number",
    "   • Logic: Unique identifier for each ticket.",
    "",
    "Assigned to",
    "   • Required column in input file: Assigned to",
    "   • Logic: Name of the person assigned to the ticket.",
    "",
    "Application",
    "   • Required column: Application Name / CI",
    "   • Logic: Application or CI related to the ticket.",
    "",
    "Tower",
    "   • Required column: Application Name / CI",
    "   • Logic: Derived from Tower_Maping.xlsx. If missing, shows 'Missing Column', 'File Missing', or 'File Error'.",
    "",
    "1. Short Description Length Check",
    "   • Required column: Short description",
    "   • Logic: Must not be empty/null and must meet minimum character length.",
    "",
    "2. Long Description Length Check",
    "   • Required column: Description",
    "   • Logic: Must not be empty/null and must meet minimum character length.",
    "",
    "3. Actual Response Time took",
    "   • Required column: Response Time",
    "   • Logic: Must be a number and in minutes.", 
    "     - Calculates time difference between 'Opened' date and first comment timestamp. Fails if response time exceeds threshold.",
    "",
    "4. Response SLA Met ?",
    "   • Required column: Response SLA",
    "   • Logic: Fail if null/None/'Breached'/True in any case; pass otherwise.",
    "",
    "5. Resolution SLA Met ?",
    "   • Required column: Resolution SLA",
    "   • Logic: Fail if null/None/'Breached'/True in any case; pass otherwise.",
    "",
    "6. KBA Tagged?",
    "    • Required column: Knowledge Article Used",
    "    • Logic: Pass if contains 'KB', 'True', or 'TRUE'; fail otherwise.",
    "",
    "7. Reopened ?",
    "    • Required column: Reopened",
    "    • Logic: Fail if null/None or value > 0; pass if value is 0.",
    "",
    "8. Work notes Length Check",
    "    • Required column: Comments and Work notes",
    "    • Logic: Must meet minimum char/word count, have punctuation, and start with a capital letter.",
            "We remove system-like headers (date/time/name/tags) from each line, then the incident’s cleaned text must be well‑formed (word_count >= 10, char_count > 50, has . , ? !, and starts with a capital) to Pass. Default character threshold is 100.",
    "    • Supported Date Formats",
    "        - YYYY-MM-DD (ISO)",
    "        - YYYY/MM/DD",
    "        - YYYY.MM.DD",
    "        - DD-MM-YYYY",
    "        - DD/MM/YYYY",
    "        - DD.MM.YYYY",
    "        - MM-DD-YYYY",
    "        - MM/DD/YYYY",
    "        - MM.DD.YYYY",
    "        - (Single-digit day/month also supported, e.g., 6/4/2022)",
    "        - Separators supported: '-', '/', '.'",
    "    • Supported Time Formats",
    "        - 24-hour: HH:MM (e.g., 09:27), HH:MM:SS (e.g., 09:27:00)",
    "        - 12-hour with AM/PM: HH:MM AM/PM, HH:MM:SS AM/PM",
    "        - Variants with dots: A.M. / P.M.",
    "        - Case-insensitive and optional spaces before AM/PM are supported",
    "    • Header Handling",
    "        - Removes optional '- Name'",
    "        - Removes optional '- Name (Work notes)' or '- Name (Additional comments)'",
    "        - If header is missing, only date/time is stripped and comment text remains",
    "",
    "9. Resolution Notes Length/ Additional comments",
    "    • Required column: Additional comments",
    "    • Logic: Must meet minimum char/word count, have punctuation, and start with a capital letter.",
    "      We remove system-like headers (date/time/name/tags) from each line, then the incident’s cleaned text must be well‑formed (10+ words (word_count &gt;= 10), char_count &gt; Additional_com_length, has . , ? !, and starts with a capital) to Pass. Default character threshold is 50.",
    "    • Supported Date Formats",
    "        - YYYY-MM-DD (ISO)",
    "        - YYYY/MM/DD",
    "        - YYYY.MM.DD",
    "        - DD-MM-YYYY",
    "        - DD/MM/YYYY",
    "        - DD.MM.YYYY",
    "        - MM-DD-YYYY",
    "        - MM/DD/YYYY",
    "        - MM.DD.YYYY",
    "        - (Single-digit day/month also supported, e.g., 6/4/2022)",
    "        - Separators supported: '-', '/'",
    "    • Supported Time Formats",
    "        - 24-hour: HH:MM (e.g., 09:27), HH:MM:SS (e.g., 09:27:00)",
    "        - 12-hour with AM/PM: HH:MM AM/PM, HH:MM:SS AM/PM",
    "        - Variants with dots: A.M. / P.M.",
    "        - Case-insensitive and optional spaces before AM/PM are supported",
    "    • Header Handling",
    "        - Removes optional '- Name'",
    "        - Removes optional '- Name (Work notes)' or '- Name (Additional comments)'",
    "        - If header is missing, only date/time is stripped and comment text remains",
    "10. Right Assignment group Usage",
    "    • Required columns: Application Name / CI, Assignment Group",
    "    • Logic: Checks if Assignment Group matches mapping in Tower_Maping.xlsx.",
    "",
    "11. Right Pending Justification Usage",
    "    • Required columns: Pending reason, Comments and Work notes, Additional comments",
    "    • Logic: Pass if Pending reason is null/None; otherwise, checks for valid justification phrases in comments.",
    "",
    "12. Related records tagged?",
    "    • Required columns: Pending reason, Related Record",
    "    • Logic: For specific pending reasons, checks if Related Record matches expected format (e.g., CHG, PRB, RITM, INC).",
    "    • For each ticket, check if the Related Record field is correctly tagged based on the Pending reason:",

    "   •   pendingchange → Related Record must contain 'CHG'",
    "   •   pendingvendor → Related Record must be present (any value except null/'None')",
    "   •   pendingproblem → must contain 'PRB'",
    "   •   pendingfulfillment → must contain 'RITM'",
    "   •   pendingincident → must contain 'INC'",
    "   •   pendingcustomer → uses external justification logic (Pending_Justification())",
    "   •   Remarks (possible results):",
    "           Pass → Related record is correctly tagged for the given pending reason.",
    "           Fail → Related record is missing or incorrectly tagged.",
    "           Fail - Missing Column → Required columns are missing from the input.",
    "           Pass → If Pending reason is null or 'None'."
    "",
    "13. Ticket Ageing Check",
    "    • Required column: Age",
    "    • Logic: Fail if Age is missing/None/above threshold; pass if below threshold.",
    "",
    "14. 3 Strike rule remainders check",
    "    • Required columns: Age, Comments and Work notes, Additional comments",
    "    • Logic: Checks for reminder phrases in comments based on ticket age.",
    "",
    "15. Reassignment check?",
    "    • Required column: Reassignment count",
    "    • Logic: Fail if count is missing/None/above threshold; pass if below threshold.",
    "",
    "16. Priority Validation",
    "    • Required columns: Priority, Impact, Urgency",
    "    • Logic: Checks if Priority matches valid combinations of Impact and Urgency.",
    "",
    "17. Category Validation",
    "    • Required column: Description and mapping file Category_Subcategory_Mapping.xlsx",
    "    • Logic: Finds the best matching keyword from the mapping based on words in the Description",
            " then checks if the ticket's Category (and Subcategory, if given) matches the mapping.",
            " Returns 'Pass' if both match, otherwise 'Fail'",
    "",
    "18. password_detected?",
    "    • Required columns: Comments and Work notes, Additional comments",
    "    • Logic: Checks for presence of password-related terms in either column.",
    "     Loop through each ticket: Combine both comment fields Split into words Check each word against the password pattern",
    "",
    "19. Closed with User Confirmation?",
    "    • Required columns: Comments and Work notes, Additional comments",
    "    • Logic: Pass if user confirmation phrases are found in either column; fail otherwise.",
    "",
    "20. Work Note Format & Content Check",
    "    • Required column: Comments and Work notes",
    "    • Logic: For each ticket’s Comments and Work notes:",
    "         - Find a date (e.g., dd/mm/yyyy, dd-mm-yyyy, or dd-MMM-yyyy).",
    "         - Take the text after that date and verify it is:",
    "         - > 30 characters and ≥ 10 words,",
    "         - Contains at least one sentence-ending punctuation (. ! ? ;), and",
    "         - The first alphabetic character after the date is uppercase.",
    "    • Remarks (possible results):",
    "         - Comprehensive → All conditions satisfied.",
    "         - Needs improvement → Date found, but content fails one or more checks.",
    "         - No match or invalid format → No recognizable date or parsing issue.",
    "         - Fail → Work note is NaN or the literal string 'None'.",
    "",
    "21. Ticket Updated Within Business Days",
    "    • Required columns: Comments and Work notes, Additional comments",
    "    • Logic: Pass if ticket was updated within the defined business days; fail otherwise.",
    "",
    "22. Process Adherence Violation Check",
    "    • Required columns: Opened, Closed, Comments and Work notes, Additional comments",
    "    • Logic: Checks if ticket updates in Comments and Work notes or Additional comments occur at least every alternate business day (excluding weekends and holidays) from Opened to Closed.",
    "    • If there is a gap of more than one business day between any two updates (including from Opened to first update and last update to Closed), the ticket fails the PA violation check; otherwise, it passes.",
    "",
    "- Strike Pattern Logics Overview:",

    "1-1-1 Pattern:",

    "Requires three consecutive business days with comments",
    "Fails if any gap is more than 1 business day",
    "Works correctly for handling duplicate timestamps",
    "2-2-1 Pattern:",

    "First two gaps can be up to 2 business days",
    "Last gap must be 1 business day",
    "Works correctly but could be optimized for clarity",
    "3-2-1 Pattern:",

    "First gap can be up to 3 business days",
    "Second gap can be up to 2 business days",
    "Last gap must be 1 business day",
    "Works correctly but could be optimized for clarity",
    "",
    "",
    "23. 3 Strike Check(1-1-1)",
    "    • Required columns: Comments and Work notes, Additional comments",
    "    • Logic: For each ticket, if its age is less than or equal to the threshold (e.g., 3 days), it is marked as 'Age <= 3 Days' and not checked further.",
    "    • If there are fewer than 4 comment timestamps, it marks 'Not Enough Comments'.",
    "    • The first three updates must be every other day, and the 4th update must be on the immediate next business day after the 3rd.",
    "    • If this pattern is followed, the ticket passes; otherwise, it fails.",
    "",
    "24. 3 Strike Check(2-2-1)",
    "    • Required columns: Comments and Work notes, Additional comments",
    "    • Logic: For tickets older than a set threshold, collect all comment timestamps from 'Comments and Work notes' and 'Additional comments'",
    "    • Check if the first four distinct comment dates follow this pattern:",
    "        - The gap between the 1st and 2nd comment is ≤ 2 business days",
    "        - The gap between the 2nd and 3rd comment is ≤ 2 business days",
    "        - The gap between the 3rd and 4th comment is ≤ 1 business day",
    "    • Remarks (possible results):",
                " - Pass → All three gaps meet the 2–2–1 rule.",
                " - Fail → At least one gap exceeds the allowed limit, or fewer than 4 distinct dates.",
                " - Age ≤ Threshold → Ticket is too new (e.g., ≤ 3 days), so the check is skipped.",
                " - Not Enough Comments → Fewer than 4 timestamps found in comments.",
    "",
    "25. 3 Strike Check(3-2-1)",
    "    • Required columns: Comments and Work notes, Additional comments",
    "    • Logic: For tickets older than a set threshold, collect all comment timestamps from 'Comments and Work notes' and 'Additional comments'.",
    "    • Check if the first four distinct comment dates follow this pattern:",
    "        - The gap between the 1st and 2nd comment is ≤ 3 business days",
    "        - The gap between the 2nd and 3rd comment is ≤ 2 business days",
    "        - The gap between the 3rd and 4th comment is ≤ 1 business day",
    "    • Remarks (possible results):",
    "        - Pass → All three gaps meet the 3–2–1 rule.",
    "        - Fail → At least one gap exceeds the allowed limit, or there are fewer than 4 distinct dates.",
    "        - Age ≤ Threshold → Ticket is too new (e.g., ≤ N days), so the check is skipped.",
    "        - Not Enough Comments → Fewer than 4 timestamps found in comments.",
    "",
    "26. Score",
    "    • Logic: Total number of validations passed for each ticket.",
    "",
    "27. Score Category",
    "    • Logic: Categorizes score: Below 75%, 75%-90%, Above 90%.",
    "",
    "28. Observations1",
    "    • Evaluates the 'Comments and Work notes' field for each ticket.",
    "    • Checks for:",
    "        - Word count (should be at least 20 words)",
    "        - Presence of punctuation",
    "        - Minimum length (at least 100 characters)",
    "        - Starts with a capital letter",
    "    • Returns feedback like 'Too few words', 'Missing punctuation', 'Too short', 'Does not start with capital' (comma-separated if multiple issues).",
    "    • If the entry is missing or 'None', marks as 'Fail'.",
    """   IMPORTANT: If any single remark in the ticket triggers a flag (e.g., "Too short", "Too few words", etc.)", then that flag is shown for the entire ticket in the Observations1 output.""",
    "       Even if most remarks are good, just one short or problematic remark will cause the flag to appear for the whole ticket.",
    "       The output does not show which specific remark triggered the flag—just that at least one did.",
    "",
    "29. Observations2",
    "    • Generates feedback based on multiple validation checks (using a feedback map).",
    "    • Maps codes to specific feedback messages:",
    "        - D: Short Description too short",
    "        - E: Long Description too short",
    "        - F: Response time too high",
    "        - G: Response SLA breached",
    "        - H: Resolution SLA breached",
    "        - I: KBA not tagged",
    "        - J: Ticket reopened",
    "        - K: worknotes is not comprehensive",
    "        - L: Resolution notes is not comprehensive",
    "        - M: Ticket is assigned to incorrect group",
    "        - N: Related record is not tagged",
    "        - O: Ticket is ageing>20 days",
    "        - P: 3 Strike rule(Confirmation) is not followed",
    "        - Q: Reassignment count is >3",
    "        - R: 1-1-1 Check is not followed",
    "        - S: 2-2-1 Check is not followed",
    "        - T: 3-2-1 Check is not followed",
    "",
#     "30. Work Notes Updated Regularly",
#     "    • Required column: Age, Comments and Work notes",
# "        • If the work note is recent (< 7 days) → it's automatically marked as Pass.",
# "        • If it's older (≥ 7 days) → it goes through a Three Strike Rule check function to decide Pass/Fail.",
# "        • If the age is missing or invalid → it's marked as Fail.",
# "        • The result is stored in the output and used for reporting.",
# " • Three Strike Rule check fuction Explained below:",
# "      • The function checks the ticket's age and comments to see if reminders were sent as per the escalation policy.",
#         """"If Age is missing or 'None' → mark as "Fail"."
#         If Age < 3 → mark as "Pass" (no reminder needed yet).
 
#         Merge 'Comments and Work notes' and 'Additional comments' into one string.
 
#         If combined comment text is empty → mark as "Fail".
#         Select reminder patterns based on age:
 
#         Age 3–5 → check for first reminder.
#         Age 6–7 → check for first + second reminders.
#         Age > 7 → check for final reminder.
#         Pattern matching:
 
#         Use regex to search for reminder phrases in the combined comment text.
#         If any match is found → "Pass", else → "Fail"."""
    "",
    "30. Has Attachments",
    "    • Required column: Has Attachments",
    "    • Logic: Check if the ticket has an attachment based on values in the specified column.",
    "    • Remarks:",
    "        - This ticket has attachment → If the value contains 'TRUE', 'true', 'True', 'YES', 'yes', or 'Yes'",
    "        - No attachment → If the value is present but doesn't match the above",
    "        - No data found → If the value is None, NaN, or the string 'None'",
    "       - Missing Column → If the 'Has Attachments' column is missing in the input file",
 ]

    
    for i, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=i+1, column=1, value=instruction)
        # Bold all validation rule names (lines starting with a number and dot)
        if instruction.strip() and instruction[0].isdigit() and '.' in instruction:
            cell.font = Font(bold=True, size=12)
        # Make required column names italic and colored to look like column headers
        elif instruction.strip().startswith('• Required column') or instruction.strip().startswith('• Required column in input file'):
            import re
            match = re.search(r':\s*(.+)', instruction)
            if match:
                col_name = match.group(1)
                # Replace column name with formatted version
                cell.value = instruction.split(':')[0] + ': ' + col_name
                cell.font = Font(italic=True, color="0070C0", size=11)
            else:
                cell.font = Font(size=11)
        else:
            cell.font = Font(size=11)
        cell.alignment = Alignment(wrap_text=True)
    
    instructions_ws.column_dimensions['A'].width = 80
 
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.download_button(
    "📥 Download Template",
    data=generate_enhanced_template(),
    file_name=f"Ticket_Audit_Template_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    help="Simple template with sample data and instructions"
)

st.markdown("###### ❇️ Instructions to fill are embedded in Template file")
st.markdown("")
st.markdown("")

st.markdown("### 1️⃣ Select Validation Rules")
 
# --- Rule selection
# List of all validation rules
all_rules = [
    "Tower", "Short Description Length Check", "Long Description Length Check",
    "Actual Response Time took", "Response SLA Met ?", "Resolution SLA Met ?", "KBA Tagged?",
    "Reopened ?", "Work notes Length Check", "Resolution Notes / Additional comment Length Check", "Right Assignment group Usage",
    "Related records tagged?", "Ticket Ageing Check", "Reassignment check?",
    "Has Attachments","Password_detected?",
    "Priority Validation", "Category Validation", "Right Pending Justification Usage",
    "Closed with User Confirmation?", #"Work Notes Updated Regularly",
    "Ticket Updated Within Business Days", "Process Adherence Violation Check",
    "3 Strike rule check(escalation policy check for Remainder)",
    "Work Note Format & Content Check", "3 Strike Check(1-1-1)","3 Strike Check(2-2-1)", "3 Strike Check(3-2-1)","Acknowledgment notes recorded in Worklog?",
    "Is the resolution summary and closure notes updated as per appropriate template?"
]
#, "3 Strike Check(1-1-1)","3 Strike Check(2-2-1)", "3 Strike Check(3-2-1)"

# # --- Select All toggle
# select_all = st.checkbox("✅ Select All Rules", value=True)

# # --- Create columns for checkboxes
# num_columns = 3
# columns = st.columns(num_columns)

# # --- Define special and normal rules
# special_rules = ["3 Strike Check(1-1-1)", "3 Strike Check(2-2-1)", "3 Strike Check(3-2-1)"]
# normal_rules = [rule for rule in all_rules if rule not in special_rules]

# # --- Collect selected rules
# selected_rules = []

# # Render normal rules as checkboxes
# for idx, rule in enumerate(normal_rules):
#     col = columns[idx % num_columns]
#     with col:
#         if st.checkbox(rule, value=select_all, key=f"rule_{idx}"):
#             selected_rules.append(rule)

# # ✅ Dropdown for 3 Strike Check (only one can be selected)
# selected_strike_rule = st.selectbox("Select 3 Strike Check Rule", options=["None"] + special_rules)

# # Add selected dropdown rule if not "None"
# if selected_strike_rule != "None":
#     selected_rules.append(selected_strike_rule)

# ✅ Now selected_rules contains all checkbox selections + the chosen dropdown rule

#################################
# MultiSelect for 3 Strike rules#
#################################
# --- Select All toggle
select_all = st.checkbox("✅ Select All Rules", value=True)

# --- Create columns for checkboxes
num_columns = 3
columns = st.columns(num_columns)

# --- Define special and normal rules
special_rules = ["3 Strike Check(1-1-1)", "3 Strike Check(2-2-1)", "3 Strike Check(3-2-1)"]
normal_rules = [rule for rule in all_rules if rule not in special_rules]

# --- Collect selected rules
selected_rules = []

# Normal rules as checkboxes
for idx, rule in enumerate(normal_rules):
    col = columns[idx % num_columns]
    with col:
        if st.checkbox(rule, value=select_all, key=f"rule_{idx}"):
            selected_rules.append(rule)

# Multi-select for 3 Strike rules
selected_strike_rules = st.multiselect(
    "Select 3 Strike Check Rule(s)",
    options=special_rules
)

# Merge both selections
selected_rules.extend(selected_strike_rules)

# --- Show selected rules (optional)
st.markdown(f"**Selected Rules ({len(selected_rules)}):**")
#st.write(selected_rules)


###########################################################
#########################################################


# --- Threshold inputs

thresholds = {}

if "Short Description Length Check" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["short_desc"] = st.number_input("Short Description Minimum Length", value=50, min_value=1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Short Description Length Logic:**
            
            This rule validates that the short description field contains sufficient detail.
            
            - **Checks the character count of the short description field**
            - **Fails if description is shorter than the specified threshold**
            - **Example:** If set to 50 characters, "Network issue" (13 chars) would fail
            - **Good example:** "Unable to connect to shared network drive from laptop" (54 chars) would pass
            """)

if "Long Description Length Check" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["long_desc"] = st.number_input("Long Description Minimum Length", value=150, min_value=1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Long Description Length Logic:**
            
            This rule ensures the long description provides comprehensive details about the issue.
            
            - **Checks the character count of the long description field**
            - **Fails if description is shorter than the specified threshold**
            - **Should include problem details, impact, and context**
            - **Example:** A 150+ character description should explain what happened, when, and how it affects the user
            """)

# Response time configuration moved to after file upload
# This allows dynamic priority selection based on uploaded data

if "Work notes Length Check" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["worknote"] = st.number_input("worknote Minimum Length threshold", value=100, min_value = 1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Work notes Length Check Logic:**
            This rule ensures that Comments and Work Notes provide meaningful and complete information.

            **Validation Steps:**

               - Remove system headers (date, time, name, and tags like Comments and Work notes).
               - Check the cleaned text for:
               - Word count ≥ 20
               - Character count > 100 (default threshold)
               - Contains punctuation (. , ? !)
               - Starts with a capital letter
                     
               **Fail Conditions:**
                - Empty or “None” value
                - Does not meet any of the above criteria
                **Purpose:** 
                     Comments should clearly explain the resolution, actions taken, and any relevant details for future reference.
            """)

if "Resolution Notes / Additional comment Length Check" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["Additional_comments_value"] = st.number_input("Resolution Notes / Additional comment Length Check", value =50, min_value= 1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Resolution Notes / Additional comment Length Check Logic:**
            This rule ensures that additional comments provide meaningful and complete information.

            **Validation Steps:**

               - Remove system headers (date, time, name, and tags like Additional comments).
               - Check the cleaned text for:
               - Word count ≥ 10
               - Character count > 50 (default threshold)
               - Contains punctuation (. , ? !)
               - Starts with a capital letter

               **Fail Conditions:**
                - Empty or “None” value
                - Does not meet any of the above criteria
            **Purpose:** Comments should clearly explain the resolution, actions taken, and any relevant details for future reference.
            """)

if "Ticket Ageing Check" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["Age Value"] = st.number_input("Enter Ageing threshold for Ticket Ageing Check",value = 20, min_value=1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Ticket Ageing Logic:**
            
            This rule identifies tickets that have been open for too long.
            
            - **Calculates the number of days between 'Opened' date and current date**
            - **Fails if ticket age exceeds the specified threshold in days**
            - **Helps identify stale or forgotten tickets**
            - **Example:** If set to 20 days, any ticket open for more than 20 days will be flagged
            """)

if "Reassignment check?" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["Reassignment threshold"] = st.number_input("Enter Reassignment threshold: ",value =3, min_value= 1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Reassignment Check Logic:**
            
            This rule monitors excessive ticket reassignments which may indicate poor initial assignment.
            
            - **Counts the number of times the 'Assigned to' field has changed**
            - **Fails if reassignment count exceeds the specified threshold**
            - **Helps identify routing issues or skill gaps**
            - **Example:** If set to 3, a ticket reassigned more than 3 times will be flagged
            """)

if "Ticket Updated Within Business Days" in selected_rules:
    col1, col2 = st.columns([2, 2])
    with col1:
        thresholds["ticket_update_days"] = st.number_input("Ticket Update Threshold (Business Days)", value=2, min_value=1, max_value=10)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Ticket Updated Within X Business Days Logic:**
            
            This rule checks if the first comment/work note is added within the specified number of business days from when the ticket was opened.
            
            - **Business days exclude weekends (Saturday and Sunday)**
            - **Example:** If set to 2 days, a ticket opened on Monday must have its first comment by Wednesday to pass
            - **Another example:** If set to 1 day, a ticket opened on Friday must have its first comment by Monday to pass
            - **The system automatically calculates business days, skipping weekends**
            """)

if '3 Strike Check(1-1-1)' in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["1-1-1 Check"] = st.number_input("Enter 3 Strike Check(1-1-1) threshold: ", value=3, min_value=1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **3 Strike Check (1-1-1) Logic:**
            
            This rule implements a progressive penalty system for tickets with multiple violations.
            
            - **Pattern:** 1st violation = warning, 2nd violation = warning, 3rd+ violation = fail
            - **Applies to tickets that exceed the threshold number of violations**
            - **Helps identify consistently problematic tickets**
            - **Example:** If set to 3, tickets with 3+ violations in any other rules will be marked as fail
            """)
        
if '3 Strike Check(2-2-1)' in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["2-2-1 Check"] = st.number_input("Enter 3 Strike Check(2-2-1) threshold: ", value=3, min_value=1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **3 Strike Check (2-2-1) Logic:**
            
            This rule implements a different progressive penalty pattern.
            
            - **Pattern:** 1st-2nd violation = warning, 3rd-4th violation = warning, 5th+ violation = fail
            - **More lenient than 1-1-1, allowing more violations before failure**
            - **Useful for complex tickets that may have legitimate challenges**
            - **Example:** If set to 3, allows up to 4 violations before marking as fail
            """)
        
if '3 Strike Check(3-2-1)' in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["3-2-1 Check"] = st.number_input("Enter 3 Strike Check(3-2-1) threshold: ", value=3, min_value=1)
    with col2:
        with st.expander("Logic"):
            st.write("""
            **3 Strike Check (3-2-1) Logic:**
            
            Validate if a ticket has updates on 4 unique business days with decreasing gaps:
            
            - **Pattern:** 1st to 2nd update ≤ 3 business days**
            - **2nd to 3rd update ≤ 2 business days**
            - **3rd to 4th update ≤ 1 business day**
            - **Example:** Updates on: Oct 1, Oct 4, Oct 6, Oct 7
            - Gaps: 3 days, 2 days, 1 day → ✅ Pass
            - If gaps are 4, 2, 1 → ❌ Fail (First gap > 3 days)
            - If only 3 update days → ❌ Fail (Not Enough Unique Days)
            """)

if "Acknowledgment notes recorded in Worklog?" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["Acknowledgment_notes_template"] = st.text_area("Enter Acknowledgment notes template to check in Worklog:", value="Thank you for reaching out to us.")
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Acknowledgment notes recorded in Worklog? Logic:**
            This rule checks if the exact acknowledgment template provided by the user
            is present in the worklog comments after cleaning timestamps and markers.
            - User can change the template anytime in the UI
            - Copy Paste your template here
            - Pass if full cleaned text matches the template
            """)

if "Is the resolution summary and closure notes updated as per appropriate template?" in selected_rules:
    col1, col2 = st.columns([3, 1])
    with col1:
        thresholds["Resolution_summary_template"] = st.text_area("Enter Resolution summary template to check:", value="Enter Resolution/Closure Summary here")
    with col2:
        with st.expander("Logic"):
            st.write("""
            **Is the resolution summary and closure notes updated as per appropriate template? Logic:**
            This rule checks if the resolution summary and closure notes match the provided template.
            - User can change the template anytime in the UI
            - Copy Paste your template here
            - Pass if full cleaned text matches the template
            """)

# if "Work Notes Updated Regularly" in selected_rules:
#     col1, col2 = st.columns([3, 1])
#     with col1:
#         thresholds["Work Notes Updated Regularly"] = st.number_input("Enter Work Notes Updated Regularly threshold: ", value=7, min_value=1, help="ℹ️ Set the number of days after which work notes are considered outdated.")
#     with col2:
#         with st.expander("Logic"):
#             st.write("""
#             **Work Notes Updated Regularly Logic:**
            
#             Required Columns: Age, Comments and Work notes, Additional comments.
                     
#             - ✅ If the **Age** of work notes is **less than the threshold**, it's marked as **Pass**.
#             - ❌ If the value is **missing** or `'None'`, it's marked as **Fail**.
#             - ⚠️ If the **Age is equal to or more than the threshold**, a **Three-Strike Rule** is applied to decide **Pass/Fail**.

# """)


# weighted scoring for your validation rules with pass fail criteria only. not for Ticket Number,Assigned to, Application, Tower, Observations1,Observations2 etc.
scoring_rules = [
    "Short Description Length Check",
    "Long Description Length Check",
    "Actual Response Time took",
    "Response SLA Met ?",
    "Resolution SLA Met ?",
    "KBA Tagged?",
    "Reopened ?",
    "Work notes Length Check",
    "Resolution Notes / Additional comment Length Check",
    "Right Assignment group Usage",
    "Related records tagged?",
    "Ticket Ageing Check",
    "Reassignment check?",
    "Priority Validation",
    "Category Validation",
    "Right Pending Justification Usage",
    "Password_detected?",
    "Closed with User Confirmation?",
    #"Work Notes Updated Regularly",
    "Ticket Updated Within Business Days",
    "Process Adherence Violation Check",
    "3 Strike rule check(escalation policy check for Remainder)",
    "Work Note Format & Content Check",
    "3 Strike Check(1-1-1)",
    "3 Strike Check(2-2-1)",
    "3 Strike Check(3-2-1)",
    "Acknowledgment notes recorded in Worklog?",
    "Is the resolution summary and closure notes updated as per appropriate template?"
]

# weights = {}
# selected_scoring_rules = [rule for rule in selected_rules if rule in scoring_rules]
# if selected_scoring_rules:
#     st.markdown("### ⚖️ Assign Weightage to Selected Rules (Optional)")
#     '''Assign a custom weight (e.g., 20%) to a specific validation rule.
#         Ensure the total weight across all selected rules sums to 100%.
#         '''
#     total_weight = 100
#     assigned_weight = 0
#     #weights = {}
#     for rule in selected_scoring_rules:
#         weight = st.number_input(f"Weight for '{rule}' (0-100):", min_value=0, max_value=100, value=0, step=5)
#         weights[rule] = weight
#         assigned_weight += weight
#     if assigned_weight > total_weight:
#         st.error(f"❌ Total assigned weight exceeds {total_weight}%. Please adjust the weights.")
#     elif assigned_weight < total_weight:
#         remaining_weight = total_weight - assigned_weight
#         unassigned_rules = [rule for rule in selected_scoring_rules if weights[rule] == 0]
#         if unassigned_rules:
#             equal_share = remaining_weight // len(unassigned_rules)
#             for rule in unassigned_rules:
#                 weights[rule] = equal_share
#             assigned_weight += equal_share * len(unassigned_rules)
#         if assigned_weight < total_weight:
#             st.warning(f"❌ Total assigned weight is less than {total_weight}%. Your Weightage score may not populate accurately")
#             st.markdown(f"**Total Assigned Weight: {assigned_weight}%**")
#     else: 
#         #assigned_weight == total_weight
#         st.success("✅ Total weight is correctly assigned to 100%.")
# else:
#     st.info("Select at least one rule with pass/fail criteria to assign weights.")
# app.py

weights = {}
selected_scoring_rules = [rule for rule in selected_rules if rule in scoring_rules]
if selected_scoring_rules:
    st.markdown("### ⚖️ Assign Weightage to Selected Rules (Optional)")
    st.markdown(
    "<div style='background-color:#ffeeba;padding:10px;border-radius:5px;'>"
    "<strong>Click below for Assign Weightage to Selected Rules</strong>"
    "</div>",
    unsafe_allow_html=True
)
    with st.expander("⚖️ **Click here 👆**"):
        st.caption("##### Assign a custom weight (e.g., 20%) to a specific validation rule. Ensure the total weight across all selected rules sums to 100%.")

        total_weight = 100
        assigned_weight = 0

        # Input weights
        for rule in selected_scoring_rules:
            weight = st.number_input(f"Weight for '{rule}' (0-100):", min_value=0, max_value=100, value=0, step=5)
            weights[rule] = weight
            assigned_weight += weight

        # Handle weight logic
        if assigned_weight > total_weight:
            st.error(f"❌ Total assigned weight exceeds {total_weight}%. Please adjust the weights.")
        elif assigned_weight < total_weight:
            if all(w == 0 for w in weights.values()):
                st.info("ℹ️ All selected rules have 0% weight. Weightage scoring will be skipped.")
            else:
                remaining_weight = total_weight - assigned_weight
                unassigned_rules = [rule for rule in selected_scoring_rules if weights[rule] == 0]
                if unassigned_rules:
                    equal_share = remaining_weight // len(unassigned_rules)
                    for rule in unassigned_rules:
                        weights[rule] = equal_share
                    assigned_weight += equal_share * len(unassigned_rules)

                if assigned_weight < total_weight:
                    st.warning(f"❌ Total assigned weight is less than {total_weight}%. Your Weightage score may not populate accurately.")
                st.markdown(f"**Total Assigned Weight: {assigned_weight}%**")
        else:
            st.success("✅ Total weight is correctly assigned to 100%.")
else:
    st.info("Select at least one rule with pass/fail criteria to assign weights.")


# --- Upload Excel

st.markdown("### 2️⃣ Upload Filled Excel File")

uploaded_file = st.file_uploader("📤 Upload Excel", type="xlsx")
 
if uploaded_file:

    input_df = pd.read_excel(uploaded_file)
    # --- Calculate Age column from Opened and Closed dates
    def parse_date(date_str):
        if pd.isnull(date_str):
            return None
        # Try multiple date formats
        date_formats = [
            '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M',
            '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M',
            '%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M',
            '%m-%d-%Y %H:%M:%S', '%m-%d-%Y %H:%M', '%m.%d.%Y %H:%M:%S', '%m.%d.%Y %H:%M',
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y', '%m/%d/%Y', '%m-%d-%Y', '%m.%d.%Y'
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(str(date_str).strip(), fmt)
            except Exception:
                continue
        return None
    
    ages = []
    for idx, row in input_df.iterrows():
        opened = parse_date(row.get('Opened'))
        closed = parse_date(row.get('Closed'))
        if opened and closed:
            age_days = (closed - opened).total_seconds() / (24*3600)
            ages.append(round(age_days))
        else:
            ages.append(None)
    input_df['Age'] = ages

    st.success("✅ File u ploaded successfully!")

    row_count = st.selectbox("Show rows (Uploaded Sheet):", [5, 10, 20, 30], index=0, key="input_row_count")
    
    st.dataframe(input_df.head(row_count), use_container_width=True)

    #st.dataframe(input_df.reset_index(drop=True), use_container_width=True)
    st.write("### Total Tickets:", len(input_df))
    
    # --- Dynamic Priority-based Response Time Configuration (after file upload)
    if "Actual Response Time took" in selected_rules:
        st.markdown("### 📊 Configure Priority-Based Response Time Thresholds")
        
        # Define priority canonicalization (same as in logic.py)
        def normalize_priority(priority_str):
            """Normalize priority to standard format: 1-Critical, 2-High, 3-Medium, 4-Low"""
            if pd.isnull(priority_str) or str(priority_str).strip() == 'None':
                return None
            
            key = str(priority_str).strip().lower()
            key = key.replace(' – ', '-').replace(' - ', '-').replace('–', '-')
            key = ' '.join(key.split())
            
            # Direct mapping
            PRIORITY_CANON_MAP = {
                '1-critical': '1-Critical', 'critical': '1-Critical', 'p1': '1-Critical',
                'severity 1': '1-Critical', 'sev1': '1-Critical', 's1': '1-Critical', 'priority 1': '1-Critical',
                '2-high': '2-High', 'high': '2-High', 'p2': '2-High',
                'severity 2': '2-High', 'sev2': '2-High', 's2': '2-High', 'priority 2': '2-High',
                '3-medium': '3-Medium', '3-moderate': '3-Medium', 'moderate': '3-Medium', 'med': '3-Medium', 'mod': '3-Medium',
                'p3': '3-Medium', 'severity 3': '3-Medium', 'sev3': '3-Medium', 's3': '3-Medium', 'priority 3': '3-Medium',
                '4-low': '4-Low', 'low': '4-Low', 'p4': '4-Low',
                'severity 4': '4-Low', 'sev4': '4-Low', 's4': '4-Low', 'priority 4': '4-Low',
            }
            
            if key in PRIORITY_CANON_MAP:
                return PRIORITY_CANON_MAP[key]
            
            # Pattern matching for variations
            import re
            for pat in (r'\bp\s*-?\s*([1-4])\b', r'\bsev(?:erity)?\s*-?\s*([1-4])\b', 
                       r'\bs\s*-?\s*([1-4])\b', r'\bpriority\s*-?\s*([1-4])\b'):
                m = re.search(pat, key)
                if m:
                    n = int(m.group(1))
                    return ['1-Critical','2-High','3-Medium','4-Low'][n-1]
            
            m = re.search(r'\b([1-4])\s*-\s*(critical|high|medium|moderate|low)\b', key)
            if m:
                word = m.group(2)
                if 'critical' in word: return '1-Critical'
                if 'high' in word:     return '2-High'
                if 'medium' in word or 'moderate' in word: return '3-Medium'
                if 'low' in word:      return '4-Low'
            
            # Keyword matching
            if 'critical' in key: return '1-Critical'
            if 'high' in key:     return '2-High'
            if 'medium' in key or 'moderate' in key: return '3-Medium'
            if 'low' in key:      return '4-Low'
            
            return None
        
        # Get unique priorities from uploaded data and normalize them
        if "Priority" in input_df.columns:
            raw_priorities = [p for p in input_df['Priority'].dropna().unique() if str(p).strip() != 'None']
            normalized_priorities = {}
            
            for p in raw_priorities:
                normalized = normalize_priority(p)
                if normalized:
                    if normalized not in normalized_priorities:
                        normalized_priorities[normalized] = str(p)
            
            unique_priorities = sorted(normalized_priorities.keys())
            
            if unique_priorities:
                priority_display = ', '.join(unique_priorities)
                st.info(f"✅ Found **{len(unique_priorities)}** priority level(s) in your data: {priority_display}")
            else:
                unique_priorities = ['1-Critical', '2-High', '3-Medium', '4-Low']
                st.warning("⚠️ Could not recognize priority values. Using default priorities.")
        else:
            unique_priorities = ['1-Critical', '2-High', '3-Medium', '4-Low']
            st.warning("⚠️ Priority column not found. Using default priorities.")
        
        # Initialize thresholds dictionary for response time
        thresholds["response_time_by_priority"] = {}
        
        st.markdown("#### Set Response Time Threshold (in minutes) for each Priority:")
        col_count = min(4, len(unique_priorities))
        cols = st.columns(col_count)
        
        # Default values for each standard priority
        default_values = {
            '1-Critical': 5,
            '2-High': 30,
            '3-Medium': 120,
            '4-Low': 1440
        }
        
        for idx, priority in enumerate(unique_priorities):
            with cols[idx % col_count]:
                default_time = default_values.get(priority, 30)
                
                time_threshold = st.number_input(
                    f"**{priority}** (minutes)",
                    value=default_time,
                    min_value=1,
                    key=f"priority_time_{priority}"
                )
                thresholds["response_time_by_priority"][priority] = time_threshold
        
        with st.expander("📖 Response Time Logic"):
            st.write("""
            **Priority-Based Response Time Logic:**
            
            This rule measures the time between ticket creation and the first response, with different thresholds per priority.
            
            - **1-Critical:** 5 minutes
            - **2-High:** 30 minutes
            - **3-Medium:** 2 hours (120 minutes)
            - **4-Low:** 24 hours (1440 minutes)
            
            **Supported Priority Formats:**
            - Numbers: 1, 2, 3, 4
            - With dashes: 1-Critical, 2-High, 3-Medium, 4-Low
            - With spaces: 1 - Critical, 2 - High, etc.
            - Short codes: P1, P2, P3, P4
            - Severity: Severity 1, Sev1, S1, etc.
            - Keywords: Critical, High, Medium, Moderate, Low
            
            - **Fails if response time exceeds the threshold for that ticket's priority**
            - **Example:** If 1-Critical is set to 5 minutes, a critical ticket with response after 6 minutes fails
            """)

    # --- 3-Strike Rule Configuration (after file upload)
    # UI for closure within X business days removed; backend uses default 3 days
    # if "3 Strike rule check(escalation policy check for Remainder)" in selected_rules:
    #     st.markdown("### 🔔 Configure 3-Strike Rule - Closure Within X Business Days")
    #     col1, col2 = st.columns([3, 1])
    #     with col1:
    #         thresholds["3_strike_closure_threshold"] = st.number_input(
    #             "Mark as PASS if ticket closed within X business days (ignores reminders):",
    #             value=3,
    #             min_value=1,
    #             max_value=10,
    #             help="Tickets closed within this many business days won't require reminder checks"
    #         )
    #     with col2:
    #         with st.expander("📖 Logic"):
    #             st.write("""
    #             **3-Strike Rule (Optimized) Logic:**
                
    #             ✅ **Pass Conditions:**
    #             - Same-day closure
    #             - Closed within X business days
    #             - Missing timestamps
    #             - Age < 3 days
    #             - Proper reminders found (based on age)
                
    #             ❌ **Fail Conditions:**
    #             - No comments found
    #             - Reminders not found when required
    #             """)
    
    # --- Process data
    st.markdown("### 3️⃣ Processed Output")
    print(selected_rules, thresholds)
    
    # --- Enhanced Pre-processing Validations with Detailed Feedback
    validation_errors = []
    validation_warnings = []
        
    # Check for required columns for each selected validation rule
    column_requirements = {
        "Short Description Length Check": ["Short description"],
        "Assigned to": ["Assigned to"],
        "Long Description Length Check": ["Description"],
        "Actual Response Time took": ["Response Time"],
        "Response SLA Met ?": ["Response SLA"],
        "Resolution SLA Met ?": ["Resolution SLA"],
        "KBA Tagged?": ["Knowledge Article Used"],
        "Reopened ?": ["Reopened"],
        "Work notes Length Check": ["Comments and Work notes"],
        "Resolution Notes / Additional comment Length Check": ["Additional comments"],
        "Right Assignment group Usage": ["Assignment Group", "Application Name / CI"],
        "Right Pending Justification Usage": ["Pending reason", "Comments and Work notes", "Additional comments"],
        "Related records tagged?": ["Pending reason", "Related Record"],
        "Ticket Ageing Check": ["Age"],
        "3 Strike rule check(escalation policy check for Remainder)": ["Age", "Comments and Work notes", "Additional comments"],
        "Reassignment check?": ["Reassignment count"],
        "Has Attachments": ["Has Attachments"],
        "Priority Validation": ["Priority", "Impact", "Urgency"],
        "Category Validation": ["Description", "Category", "Subcategory"],
        "Password_detected?": ["Comments and Work notes", "Additional comments"],
        "Closed with User Confirmation?": ["Comments and Work notes", "Additional comments"],
        #"Work Notes Updated Regularly": ["Age", "Comments and Work notes", "Additional comments"],
        "Ticket Updated Within Business Days": ["Opened", "Comments and Work notes", "Additional comments"],
        "Process Adherence Violation Check": ["Opened", "Closed", "Comments and Work notes", "Additional comments"],
        "3 Strike Check(1-1-1)": ["Age", "Comments and Work notes", "Additional comments"],
        "3 Strike Check(2-2-1)": ["Age", "Comments and Work notes", "Additional comments"],
        "3 Strike Check(3-2-1)": ["Age", "Comments and Work notes", "Additional comments"],
        "Observations1": ["Comments and Work notes"],
        "Observation2": [], # This depends on other columns, handled in logic
        "Acknowledgment notes recorded in Worklog?": ["Comments and Work notes", "Additional comments"],
        "Is the resolution summary and closure notes updated as per appropriate template?": ["Comments and Work notes", "Additional comments"]

    }
    
    missing_columns = []
    for rule in selected_rules:
        if rule in column_requirements:
            required_cols = column_requirements[rule]
            for required_col in required_cols:
                if required_col not in input_df.columns:
                    missing_columns.append({
                        'rule': rule,
                        'required_column': required_col
                    })
    
    # Display missing column warnings/errors
    if missing_columns:
        st.warning("⚠️ **Error: The following columns are missing in input file, Please add the missing columns or uncheck the validation rules above.**")
        # Group missing columns by rule
        from collections import defaultdict
        rule_to_columns = defaultdict(list)
        for missing in missing_columns:
            rule_to_columns[missing['rule']].append(missing['required_column'])
        for rule, cols in rule_to_columns.items():
            st.error(f"❌ **{rule}** requires column(s): {', '.join(f'`{col}`' for col in cols)}")
        # Show suggestions for similar columns
        # for missing in missing_columns:
        #     suggestions = suggest_similar_columns(input_df, missing['required_column'])
        #     if suggestions:
        #         st.info(f"💡 Similar columns found for `{missing['required_column']}`: {', '.join([f'`{s}`' for s in suggestions])}")
        # Allow user to choose whether to continue
        continue_anyway = st.checkbox(
            f"⚠️ Continue anyway ({len(missing_columns)} rule(s) will FAIL for all tickets)",
            help="All tickets will show 'Fail - Missing Column' for the affected validation rules."
        )
        
        if not continue_anyway:
            #st.info("📋 **Please add the missing columns or uncheck the validation rules above.**")
            st.stop()
        else:
            st.warning(f"⚠️ **Proceeding with {len(missing_columns)} missing column(s).**")

    # Tower validation logic (existing code)
    if "Tower" in selected_rules:
        # Check 1: Required input file column validation
        required_column = "Application Name / CI"
        if required_column not in input_df.columns:
            # Collect all required columns for selected rules
            required_cols = [col for rule in selected_rules if rule in column_requirements for col in column_requirements[rule]]
            missing_cols = [col for col in required_cols if col not in input_df.columns]
            st.error("❌ **Tower Logic Error: Missing Required Column(s)**")
            st.markdown(f"""
            **To run the selected validation rules, you need to add the following column(s):** `{', '.join(missing_cols)}` **in your input Excel file**
            
            **📋 Current available columns in your file:**
            {list(input_df.columns)}
            
            **✅ Please:**
            1. Add the missing column(s) to your Excel file: `{', '.join(missing_cols)}`
            2. Re-upload the file with the required columns
            """)
            st.stop()
        
        # Check 2: Tower mapping file validation
        import os
        tower_mapping_file = "Tower_Maping.xlsx"
        if not os.path.exists(tower_mapping_file):
            st.error("❌ **Tower Logic Error: Missing Tower Mapping File**")
            st.markdown(f"""
            **To run Tower logic, you need to add the file:** `{tower_mapping_file}` **in the current folder**
            
            **📁 Current folder location:** `{os.getcwd()}`
            
            **✅ Please:**
            1. Add the `{tower_mapping_file}` file to this folder
            2. Ensure the file contains columns: `Application Name` and `Tower`
            3. Re-run the analysis
            """)
            st.stop()
    
        # Check 3: Tower mapping file structure validation
        try:
            tower_df = pd.read_excel(tower_mapping_file)
            required_tower_columns = ["Application Name", "Tower"]
            missing_tower_columns = [col for col in required_tower_columns if col not in tower_df.columns]
            
            if missing_tower_columns:
                st.error("❌ **Tower Logic Error: Invalid Tower Mapping File**")
                st.markdown(f"""
                **The Tower mapping file `{tower_mapping_file}` is missing required columns:** `{missing_tower_columns}`
                
                **📋 Current columns in Tower file:**
                {list(tower_df.columns)}
                
                **✅ Please ensure your Tower mapping file contains:**
                - `Application Name` column (matching your input data)
                - `Tower` column (containing tower/business unit names)
                """)
                st.stop()
            
            # Check 4: Application coverage validation (only show warning if there are issues)
            input_applications = set(input_df[required_column].dropna().unique())
            tower_applications = set(tower_df['Application Name'].dropna().unique())
            missing_applications = input_applications - tower_applications
            
            if missing_applications:
                missing_list = ', '.join(list(missing_applications)[:5])
                more_text = f" and {len(missing_applications)-5} more" if len(missing_applications) > 5 else ""
                st.warning(f"⚠️ **Tower Mapping Warning:** Some applications not found in tower mapping: {missing_list}{more_text}")
                st.info("💡 Processing will continue with 'Unknown Tower' for unmapped applications")
                
        except Exception as e:
            st.error(f"❌ **Tower Logic Error:** Cannot read tower mapping file: {str(e)}")
            st.stop()
    
    try:
        processed_df = cached_process_uploaded_file(input_df, selected_rules, thresholds,weights)
        row_count_out = st.selectbox("Show rows (Processed Output):", [5, 10, 20, 30], index=0, key="output_row_count")
        st.dataframe(processed_df.head(row_count_out), use_container_width=True)
        #st.dataframe(processed_df.reset_index(drop=True), use_container_width=True)
    except ValueError as ve:
        if "Length of values" in str(ve):
            if ("Tower" in selected_rules) or ("Right Assignment group Usage" in selected_rules):
                st.error("🔴 Tower Mapping Error: Mismatch between data and tower mapping file")
                st.error("❌ This usually happens when Application names in your data don't match the Tower_Mapping.xlsx file")
                st.info("💡 Suggestions:")
                st.info("• Check if 'Tower_Maping.xlsx' file exists in the same folder")
                st.info("• Verify Application names match between your data and tower mapping file")  
                st.info("• Or remove 'Tower' from validation rules to continue without tower mapping")
            else:
                st.error("🔴 Data Processing Error: Mismatch between data and output columns. This may be caused by a bug in one of the selected validation rules.")
                st.info("💡 Suggestions:")
                st.info("• Check if your input data matches the requirements for the selected rules")
                st.info("• If the problem persists, contact the tool maintainer with your input file and selected rules")
            st.stop()
        else:
            st.error(f"🔴 Data Processing Error: {str(ve)}")
            st.stop()
    except Exception as e:
        st.error(f"🔴 Analysis failed: {str(e)}")
        st.stop()
 
    # --- Enhanced Download Results
    def to_enhanced_excel(df):
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main results sheet
            df.to_excel(writer, index=False, sheet_name='Audit Results')
            workbook = writer.book
            worksheet = writer.sheets['Audit Results']
            
            # Style the headers
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')  # Professional blue
            header_alignment = Alignment(horizontal='left', vertical='bottom')  # Left and bottom alignment
            
            # Apply header styling
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Color only the Score data cells (not the header)
            if 'Score' in df.columns:
                score_col = df.columns.get_loc('Score') + 1
                for row in range(2, len(df) + 2):  # Start from row 2 to skip header
                    cell = worksheet.cell(row=row, column=score_col)
                    if cell.value is not None:
                        # Make only Score data cells green (not the header)
                        cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Green
            
            # Create summary sheet
            if 'Score' in df.columns:
                summary_data = {
                    'Metric': ['Total Records', 'Average Score', 'Excellent (≥75%)', 'Good (50-74%)', 'Needs Improvement (<50%)'],
                    'Value': [
                        len(df),
                        f"{df['Score'].mean():.1f}%",
                        len(df[df['Score'] >= 75]),
                        len(df[(df['Score'] >= 50) & (df['Score'] < 75)]),
                        len(df[df['Score'] < 50])
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, index=False, sheet_name='Summary')
                
                # Style summary sheet
                summary_ws = writer.sheets['Summary']
                for col in range(1, 3):
                    cell = summary_ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='left', vertical='bottom')  # Match template alignment
                
                summary_ws.column_dimensions['A'].width = 25
                summary_ws.column_dimensions['B'].width = 15

        output.seek(0)
        return output
    
    st.download_button(
        label="📥 Download Enhanced Results",
        data=to_enhanced_excel(processed_df),
        file_name=f"Audit_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Download professionally formatted results with color-coded scores and summary sheet"
    )

    ################################
    ## --- Dashboard Visualization##
    ################################
    if st.checkbox("📈 Show Dashboard"):
        from streamlit_echarts import st_echarts  # st_echarts renders ECharts option dict [1](https://streamlit.io/components?category=charts)[2](https://echarts.streamlit.app/)

        st.subheader("📊 Overall Average Audit Score")

        # --- Ensure numeric + safe KPI ---
        processed_df["Score"] = pd.to_numeric(processed_df["Score"], errors="coerce")
        avg_score = float(processed_df["Score"].fillna(0).mean())

        # Clamp and round for stable display
        avg_score = max(0.0, min(100.0, avg_score))
        avg_score_rounded = round(avg_score, 1)

        # --- ECharts Gauge (Primary) ---
        #st.subheader("📊 ECharts Gauge")

        gauge_option = {
            "tooltip": {"formatter": "{a}<br/>{b}: {c}"},
            "series": [
                {
                    "name": "Average Audit Score",
                    "type": "gauge",
                    "startAngle": 180,
                    "endAngle": 0,
                    "min": 0,
                    "max": 100,
                    "splitNumber": 5,
                    # Put gauge center slightly lower to look like a semicircle
                    "center": ["50%", "65%"],
                    "radius": "95%",

                    # ✅ Correct thresholds: 0–60 red, 60–80 yellow, 80–100 green
                    "axisLine": {
                        "lineStyle": {
                            "width": 35,
                            "color": [
                                [0.60, "#FF4C4C"],   # Red 0-60
                                [0.80, "#FFFF00"],   # Yellow 60-80
                                [1.00, "#32CD32"],   # Green 80-100
                            ],
                        }
                    },

                    "axisTick": {"length": 8, "lineStyle": {"color": "auto", "width": 2}},
                    "splitLine": {"length": 14, "lineStyle": {"color": "auto", "width": 3}},
                    "axisLabel": {"distance": 18, "color": "#333", "fontSize": 17,"fontWeight":"bold"},

                    # Pointer: use default (more reliable than rect); keep consistent sizing
                    "pointer": {"length": "80%", "width": 8},
                    "anchor": {"show": True, "showAbove": True, "size": 10},

                    # Title + value styling
                    "title": {"show": True, "offsetCenter": [0, "-10%"], "fontSize": 14},
                    "detail": {
                        "valueAnimation": True,
                        "formatter": "{value}",
                        "fontSize": 60,
                        "offsetCenter": [0, "35%"],
                        "color": "navy",
                    },

                    "data": [{"value": avg_score_rounded,"name":"Audit Score"}],
                    "title": {"show": True, "offsetCenter": [0, "-30%"], "text": "Overall Audit Score (Average)"}
                }
            ],
        }

        # key helps Streamlit reuse component cleanly on reruns
        st_echarts(gauge_option, height="550px", key="audit_score_echarts")

        
        #2 Score Range Distribution

        import plotly.express as px
        st.subheader("📊 Score Range Distribution")
        score_cat_counts = processed_df['Score Category'].value_counts().reset_index()
        score_cat_counts.columns = ['Score Category', 'Count']

        fig_score_cat = px.pie(
            score_cat_counts,
            names='Score Category',
            values='Count',
            title='Score Range Distribution'
        )

        st.plotly_chart(fig_score_cat, use_container_width=True)



# Audit Score Tower Wise(Average Audit Score)
# What Is Being Calculated?
# Audit Score: A numerical value assigned to each ticket based on compliance with audit criteria.
# Tower: Represents a business unit or functional area.
# The code calculates the mean audit score for each tower.
        if 'Tower' in processed_df.columns and 'Score' in processed_df.columns:
            import plotly.express as px
            tower_avg = processed_df.groupby('Tower')['Score'].mean().reset_index()
            tower_avg['Score'] = tower_avg['Score'].round().astype(int)  # Round to remove decimals

            fig_tower_avg = px.bar(
                tower_avg,
                x='Tower',
                y='Score',  # This must match the actual column name
                title='Audit Score Tower Wise(Average Audit Score)',
                text='Score',  # Also must match the actual column name
                color='Tower'
            )

            fig_tower_avg.update_traces(textposition='outside')

            max_score = tower_avg['Score'].max()
            fig_tower_avg.update_layout(
                yaxis=dict(title='Average Audit Score', range=[0, max_score + 5]),  # Custom y-axis label
                uniformtext_minsize=8,
                uniformtext_mode='hide'
            )
            st.plotly_chart(fig_tower_avg, use_container_width=True)



 
        # Application-wise Audit Score
        validation_cols = [
                "Short Description", "Long Description", "Response Time",
                "Response SLA Met", "Resolution SLA Met", "KBA Tagged?", "Reopened?",
                "Work notes Length", "Resolution Notes Length", "Assignment group check", 
                "Related records tagged?",
                "Ticket Ageing Check", "3 Strike rule remainders check", "Reassignment check?",
                "Priority Validation", "Pending Justification",
                "Closed with User Confirmation?", "Work Notes Updated Regularly",
                "password_detected?", "Ticket Updated Within Business Days", "PA violation Check"
            ]

        if 'Application' in processed_df.columns:
            app_avg = processed_df.groupby('Application')['Score'].mean().reset_index()
            app_avg['Score'] = app_avg['Score'].round().astype(int)  # Round to remove decimals
            fig_app_avg = px.bar(app_avg, x='Application', y='Score', title='Audit Score Application Wise',text='Score',color = 'Application')
    
            

            st.plotly_chart(fig_app_avg, use_container_width=True)
#################################

##################################
        # Category-wise Pass %
        validation_cols = [
                "Short Description", "Long Description", "Response Time",
                "Response SLA Met", "Resolution SLA Met", "KBA Tagged?", "Reopened?",
                "Work notes Length Check", "Resolution Notes Length", "Assignment group check", "Related records tagged?",
                "Ticket Ageing Check", "3 Strike rule remainders check", "Reassignment check?",
                "Priority Validation","Category Validation", "Pending Justification",
                "Closed with User Confirmation?",
                "password_detected?", "Ticket Updated Within Business Days", "PA violation Check"
            ]
        
        category_stats = []
        for col in validation_cols:
            if col in processed_df.columns:
                total = processed_df[col].count()
                passed = processed_df[col].str.lower().eq('pass').sum()
                percent = round((passed / total) * 100, 2) if total > 0 else 0
                category_stats.append((col, percent, passed))
            else:
                category_stats.append((col, 0, 0))  # Include missing columns with 0%

        cat_df = pd.DataFrame(category_stats, columns=['Category', 'Pass %', 'Pass Count'])
        cat_df = cat_df.sort_values('Pass %')

        # Plotly bar chart with unique colors and visible labels
        fig3 = px.bar(
            cat_df,
            x='Pass %',
            y='Category',
            orientation='h',
            title="Category-wise Pass% Vs Count",
            text='Pass Count',
            color='Category',  # Assign different color to each bar
            color_discrete_sequence=px.colors.qualitative.Safe  # Use a safe color palette
        )

        fig3.update_traces(textposition='outside')

        # Ensure all category names are visible and layout is clean
        fig3.update_layout(
            yaxis=dict(
                tickmode='array',
                tickvals=cat_df['Category'],
                ticktext=cat_df['Category']
            ),
            margin=dict(l=150, r=20, t=50, b=50),  # Extra left margin for long labels
            height=600
        )

        st.plotly_chart(fig3, use_container_width=True)


        #######3
        # Enhanced Associate Performance Chart - Shows both ticket count and average score

        if 'Assigned to' in processed_df.columns and 'Score' in processed_df.columns:
            import plotly.express as px
            import plotly.graph_objects as go

            # Calculate both ticket count and average score per associate
            assigned_counts = processed_df['Assigned to'].value_counts().reset_index()
            assigned_counts.columns = ['Assigned to', 'Ticket Count']
            
            # Calculate average scores per associate
            assigned_scores = processed_df.groupby('Assigned to')['Score'].mean().reset_index()
            assigned_scores['Avg Score'] = assigned_scores['Score'].round(1)
            
            # Merge the data
            assigned_data = assigned_counts.merge(assigned_scores, on='Assigned to')
            
            # Create custom text showing both metrics clearly
            assigned_data['Display Text'] = assigned_data.apply(
                lambda row: f"{int(row['Ticket Count'])} tickets | {row['Avg Score']:.1f}%", axis=1
            )
            
            # Create color scale based on average score (red to green gradient)
            # Normalize scores to 0-1 range for color mapping
            min_score = assigned_data['Avg Score'].min()
            max_score = assigned_data['Avg Score'].max()
            if max_score > min_score:
                assigned_data['Color Value'] = (assigned_data['Avg Score'] - min_score) / (max_score - min_score)
            else:
                assigned_data['Color Value'] = 0.5

            fig_assigned = px.bar(
                assigned_data,
                x='Ticket Count',
                y='Assigned to',
                orientation='h',
                title="Associate Performance: Ticket Count & Average Score",
                text='Display Text',
                color='Avg Score',
                color_continuous_scale='RdYlGn',  # Red to Yellow to Green
                hover_data=['Avg Score']
            )

            # Enhanced styling with smoother appearance
            fig_assigned.update_traces(
                textposition='outside',
                textfont=dict(size=12, color='black', family='Arial'),
                marker=dict(line=dict(width=0.3, color='rgba(0,0,0,0.2)'))
            )

            max_count = assigned_data['Ticket Count'].max()
            fig_assigned.update_layout(
                yaxis=dict(
                    title=dict(text='Assigned To', font=dict(size=13, color='#333333'))
                ),
                xaxis=dict(
                    title=dict(text='Number of Tickets', font=dict(size=13, color='#333333')), 
                    range=[0, max_count + 8]
                ),
                title=dict(
                    text="Associate Performance: Ticket Count & Average Score",
                    font=dict(size=15, color='#333333')
                ),
                uniformtext_minsize=10,
                uniformtext_mode='hide',
                height=600,
                coloraxis_colorbar=dict(
                    title=dict(text="Average Score (%)", font=dict(size=11, color='#333333')),
                    borderwidth=0
                ),
                plot_bgcolor='white',
                paper_bgcolor='white',
                margin=dict(l=150, r=50, t=80, b=50)
            )

            st.plotly_chart(fig_assigned, use_container_width=True)

        ########
        # Observations Summary
        if 'Observations1' in processed_df.columns:
            st.subheader("📝 Observations Summary")
            # Split all comma-separated feedbacks, flatten, and count each unique parameter
            from collections import Counter
            all_params = []
            for obs in processed_df['Observations1'].dropna():
                all_params.extend([param.strip() for param in obs.split(',') if param.strip()])
            param_counts = Counter(all_params)
            obs_summary_df = pd.DataFrame(param_counts.items(), columns=["Parameter", "Count"]).sort_values("Count", ascending=False)
            st.table(obs_summary_df)
        # Observations2 Summary
        if 'Observations2' in processed_df.columns:
            st.subheader("📝 Observations Summary")
            st.write(processed_df['Observations2'].value_counts())
